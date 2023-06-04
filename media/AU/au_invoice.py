import datetime
from django.http import HttpResponse
from openpyxl import load_workbook


def make_au_invoice(queryset):
    # with load_workbook("D:\\My data\\DivyaPrakash\\media\\AU\\121.AU-Nov -Feb2022-23 - Copy.xlsx") as wb:
    wb = load_workbook("D:\\My data\\DivyaPrakash\\media\\AU\\121.AU-Nov -Feb2022-23 - Copy.xlsx")
    ws = wb.active
    next_index = 0
    legal_sum = 0
    vetting_sum = 0
    total_sum = 0
    for query in queryset:
        index = queryset.index(query) + 1
        print(index)
        ws['A' + str(index + 1)] = index
        ws['B' + str(index + 1)] = datetime.date.today()
        ws['C' + str(index + 1)] = query['customer_name']
        ws['D' + str(index + 1)] = query['location']
        ws['E' + str(index + 1)] = query['rlos_number']
        ws['F' + str(index + 1)] = query['lead_number']
        ws['G' + str(index + 1)] = query['lan_number']
        ws['H' + str(index + 1)] = query['department_type']
        ws['I' + str(index + 1)] = query['reject_hold']
        ws['O' + str(index + 1)] = query['remarks']
        print(query['price'])
        print(type(query['price']))
        if query['type'] == 'legal':
            ws['J' + str(index + 1)] = query['price']
            ws['K' + str(index + 1)] = 'NA'
            ws['N' + str(index + 1)] = query['price']
            legal_sum += int(query['price'])
        elif query['type'] == 'vetting':
            ws['K' + str(index + 1)] = query['price']
            ws['J' + str(index + 1)] = 'NA'
            ws['N' + str(index + 1)] = query['price']
            vetting_sum += int(query['price'])
        total_sum += query['price']
        ws['L' + str(index + 1)] = 'NA'
        ws['M' + str(index + 1)] = 'NA'
        next_index = len(queryset) + 2
    ws['C' + str(next_index)] = 'TOTAL'
    ws['J' + str(next_index)] = str(legal_sum)
    ws['K' + str(next_index)] = vetting_sum
    ws['L' + str(next_index)] = 4000
    ws['N' + str(next_index)] = total_sum
    print(legal_sum)
    print(vetting_sum)
    print(total_sum)
    print(ws['N' + str(next_index)].value)
    response = HttpResponse(content_type='text/xlsx')
    response['Content-Disposition'] = 'attachment; filename=AU_{}-{}.xlsx'.format(queryset[0]['date'].month,
                                                                                  queryset[0]['date'].year)
    wb.save(response)
    return response
