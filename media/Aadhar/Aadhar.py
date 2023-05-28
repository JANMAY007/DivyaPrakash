from openpyxl import load_workbook
from datetime import datetime

def make_aadhar_invoice(data, queryset):
    wb = load_workbook('1.Aadhar Gaurang N Gandhi APR-22 - Copy.xlsx')
    ws = wb.active
    ws['C4'] = data['name'] + ' ' + data['address_line_1'] + '/n' + data['address_line_2']
    ws['C8'] = data['pan_number']
    ws['C9'] = data['invoice_number']
    ws['C10'] = datetime.now().strftime('%d-%m-%Y')
    ws['C11'] = data['bank_name'] + '\n' + data['bank_account_number']
    ws['C12'] = data['bank_ifsc_code']
    ws['C14'] = queryset[0]['party_address']
    ws['C19'] = queryset[0]['delivery_address']

    wb.save('1.Aadhar Gaurang N Gandhi APR-22 - Copy.xlsx')
