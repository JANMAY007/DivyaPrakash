from openpyxl import load_workbook
from datetime import datetime
import calendar

def make_aadhar_invoice(data, queryset):
    wb = load_workbook('D:\\My data\\DivyaPrakash\\media\\Aadhar\\1.Aadhar Gaurang N Gandhi APR-22 - Copy.xlsx')
    ws = wb.active
    ws['C4'] = data['name'] + ' ' + data['address_line_1'] + '/n' + data['address_line_2']
    ws['C8'] = data['pan_number']
    ws['C9'] = data['invoice_number']
    ws['C10'] = datetime.now().strftime('%d-%m-%Y')
    ws['C11'] = data['bank_name'] + '\n' + data['bank_account_number']
    ws['C12'] = data['bank_ifsc_code']
    ws['C14'] = queryset[0]['party_address']
    ws['C19'] = queryset[0]['delivery_address']
    index = 0
    queryset = list(queryset)
    for query in queryset:
        index = queryset.index(query) + 25
        ws['A' + str(index)] = index - 25
        ws['B' + str(index)] = query['application_number']
        ws['C' + str(index)] = query['customer_name']
        ws['D' + str(index)] = query['branch']
        ws['E' + str(index)] = calendar.month_abbr[query['date'].month] + "'" + str(query['date'].year)[2:]
        if query['type'] == 'vetting':
            ws['G' + str(index)] = query['price']
        else:
            ws['F' + str(index)] = query['price']
        ws['H' + str(index)] = '=SUM(F' + str(index) + ':G' + str(index) + ')'
    ws['F' + str(index + 1)] = '=SUM(F25:F' + str(index) + ')'
    ws['G' + str(index + 1)] = '=SUM(G25:G' + str(index) + ')'
    ws['H' + str(index + 1)] = '=SUM(F' + str(index + 1) + ':G' + str(index + 1) + ')'
    wb.save('D:\\My data\\DivyaPrakash\\media\\Aadhar\\1.Aadhar Gaurang N Gandhi APR-22 - Copy.xlsx')
